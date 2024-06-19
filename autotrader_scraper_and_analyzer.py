import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import cv2
import numpy as np
from google.cloud import vision
from google.oauth2 import service_account
import openpyxl
import pandas as pd
import re
import os
import time
from collections import Counter

class AutoTraderScraper:
    def __init__(self, service_account_file, chrome_driver_path):
        self.credentials = service_account.Credentials.from_service_account_file(service_account_file)
        self.client = vision.ImageAnnotatorClient(credentials=self.credentials)
        self.chrome_driver_path = chrome_driver_path
        self.button_clicked = False

    def download_image(self, image_url):
        response = requests.get(image_url, stream=True)
        if response.status_code == 200:
            image_data = np.asarray(bytearray(response.content), dtype="uint8")
            image = cv2.imdecode(image_data, cv2.IMREAD_COLOR)
            return image
        return None

    def detect_number_plate_region(self, image, idx):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.bilateralFilter(gray, 11, 17, 17)
        edged = cv2.Canny(gray, 30, 200)
        contours, _ = cv2.findContours(edged.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        contours = sorted(contours, key=cv2.contourArea, reverse=True)[:10]

        if not os.path.exists('plate_images'):
            os.makedirs('plate_images')

        for contour in contours:
            approx = cv2.approxPolyDP(contour, 10, True)
            if len(approx) == 4:
                x, y, w, h = cv2.boundingRect(approx)
                plate_image = image[y:y + h, x:x + w]
                cv2.imwrite(os.path.join('plate_images', f'number_plate_{idx}.jpg'), plate_image)
                return plate_image
        return None

    def detect_uk_number_plates(self, text):
        uk_plate_pattern = r'\b([A-Z]{2}[0-9]{2} [A-Z]{3})\b'
        matches = re.findall(uk_plate_pattern, text)
        return matches

    def preprocess_image(self, image):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.bilateralFilter(gray, 11, 17, 17)
        return gray

    def detect_number_plate_google(self, image):
        _, encoded_image = cv2.imencode('.jpg', image)
        content = encoded_image.tobytes()
        image = vision.Image(content=content)
        response = self.client.text_detection(image=image)
        texts = response.text_annotations
        if texts:
            detected_text = texts[0].description.strip()
            return self.detect_uk_number_plates(detected_text)
        return []

    def fetch_car_details(self, driver, link):
        driver.get(link)
        wait = WebDriverWait(driver, 2)

        if not self.button_clicked:
            self._click_accept_button(wait, driver)

        driver.switch_to.default_content()

        car_name = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h1[data-gui="advert-title"]'))).text.strip()
        price = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h2[data-testid="advert-price"]'))).text.strip()
        car_type = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'p[data-testid="advert-subtitle"]'))).text.strip()
        mileage = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'Mileage')]/following-sibling::p"))).text.strip()
        registration = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'Registration')]/following-sibling::p"))).text.strip()

        seller, location = self._fetch_seller_and_location(driver)

        number_plate = self._extract_image_urls_and_detect_number_plate(wait, driver)

        return car_name, car_type, price, mileage, registration, seller, location, number_plate

    def _click_accept_button(self, wait, driver):
        try:
            accept_button = wait.until(EC.presence_of_element_located((By.XPATH, '//button[@title="Accept All"]')))
            driver.execute_script("arguments[0].click();", accept_button)
            self.button_clicked = True
        except:
            self._check_iframes_for_accept_button(wait, driver)

    def _check_iframes_for_accept_button(self, wait, driver):
        try:
            iframes = driver.find_elements(By.TAG_NAME, "iframe")
            for iframe in iframes:
                driver.switch_to.frame(iframe)
                try:
                    accept_button = wait.until(EC.presence_of_element_located((By.XPATH, '//button[@title="Accept All"]')))
                    driver.execute_script("arguments[0].click();", accept_button)
                    self.button_clicked = True
                    break
                except:
                    driver.switch_to.default_content()
        except Exception as e:
            print(f"Error while switching to iframe: {e}")

    def _fetch_seller_and_location(self, driver):
        try:
            seller = driver.find_element(By.XPATH, "//section[@data-testid='advert-seller-details']//span[contains(@class, 'at__sc-') and not(contains(@class, 'sc-wi62cf-'))]").text.strip()
            if seller.endswith("Find out more"):
                seller = seller.rsplit("Find out more", 1)[0].strip()
        except:
            seller = "N/A"

        try:
            section = driver.find_element(By.XPATH, "//section[@data-testid='advert-seller-details']")
            section_text = section.text.strip()
            if "miles away" in section_text:
                for line in section_text.split('\n'):
                    if "miles away" in line:
                        location = ' '.join(line.split()[:-3])
                        break
            else:
                location = "N/A"
        except Exception as e:
            location = "N/A"
            print(f"An error occurred: {e}")

        return seller, location

    def _extract_image_urls_and_detect_number_plate(self, wait, driver):
        number_plates = []
        try:
            view_gallery_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="gallery-view-more-button"]')))
            view_gallery_button.click()

            image_grid = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[data-testid="image-grid-component"]')))
            image_tags = image_grid.find_elements(By.CSS_SELECTOR, 'img.at__sc-2lrp52-1.dVCMUy.atds-image')
            image_urls = [img.get_attribute('src') for img in image_tags]

            for idx, image_url in enumerate(image_urls):
                image = self.download_image(image_url)
                if image is not None:
                    plate_region = self.detect_number_plate_region(image, idx + 1)
                    if plate_region is not None:
                        preprocessed_image = self.preprocess_image(plate_region)
                        number_plate_google = self.detect_number_plate_google(preprocessed_image)
                        number_plates.extend(number_plate_google)
        except Exception as e:
            print(f"Error extracting image URLs: {e}")

        if len(number_plates) >= 2:
            most_common_plate, count = Counter(number_plates).most_common(1)[0]
            if count >= 2:
                return most_common_plate
        return "No number plate detected"

    def save_links_to_spreadsheet(self, links, filename='autotrader_uk_details.xlsx'):
        options = Options()
        options.headless = False
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        service = Service(self.chrome_driver_path)
        driver = webdriver.Chrome(service=service, options=options)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'AutoTrader UK Details'
        self._write_headers(sheet)
        workbook.save(filename)

        for idx, link in enumerate(links, start=2):
            car_details = self.fetch_car_details(driver, link)
            self._write_to_sheet(workbook, sheet, link, car_details, idx)
            workbook.save(filename)

        driver.quit()

    def _write_headers(self, sheet):
        headers = ['AutoTrader Link', 'Car Name', 'Type of Car', 'Price', 'Mileage', 'Registration Year', 'Seller', 'Location', 'Number Plate', 'MOT Expiry', 'MOT History', 'Price Score', 'Mileage Score', 'Year Score', 'MOT Score', 'Total Score']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

    def _write_to_sheet(self, workbook, sheet, link, car_details, idx):
        sheet[f'A{idx}'] = link
        for col, detail in enumerate(car_details, start=2):
            sheet.cell(row=idx, column=col, value=detail)

    @staticmethod
    def read_links_from_file(filename):
        car_id_pattern = re.compile(r'car-details/(\d+)')
        unique_ids = set()
        unique_links = []

        with open(filename, 'r') as file:
            for line in file:
                link = line.strip()
                if link:
                    match = car_id_pattern.search(link)
                    if match:
                        car_id = match.group(1)
                        if car_id not in unique_ids:
                            unique_ids.add(car_id)
                            unique_links.append(link)

        return unique_links

    def get_car_details_and_mot_history(self, driver, plate):
        url = f"https://www.carcheckfree.co.uk/cardetails/{plate}"
        
        def retry_request():
            retries = 5
            delay = 100
            for attempt in range(retries):
                try:
                    car_details = {}

                    # Attempt to load the page
                    driver.get(url)

                    # Retrieve MOT expiry text
                    mot_expiry_text = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, 'mot-expiry-text'))
                    ).text
                    car_details['mot_expiry'] = mot_expiry_text.replace('Expires: ', '')

                    # Click on view MOT button
                    view_mot_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//div[@class="freechecksection shadow seefullmothistory showspinner"]//a'))
                    )
                    view_mot_button.click()
                    
                    time.sleep(10) # Evade rate limiting

                    # Retrieve MOT history
                    mot_history = []
                    mot_history_elements = WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CLASS_NAME, 'table-main'))
                    )

                    for element in mot_history_elements:
                        test_date = element.find_element(By.CLASS_NAME, 'testdate').text
                        mileage = element.find_element(By.CLASS_NAME, 'mileagenumber').text

                        try:
                            expiry_date = element.find_elements(By.CLASS_NAME, 'mileagenumber')[1].text
                        except IndexError:
                            expiry_date = "N/A"

                        comments = element.find_elements(By.CLASS_NAME, 'commentsp')
                        comments_text = [comment.text for comment in comments]

                        mot_history.append({
                            'test_date': test_date,
                            'mileage': mileage,
                            'expiry_date': expiry_date,
                            'comments': comments_text
                        })

                    car_details['mot_history'] = mot_history

                    # If all steps are successful, return the car details
                    return car_details

                except Exception as e:
                    print(f"Attempt {attempt + 1} failed: {e}")
                    if attempt < retries - 1:
                        time.sleep(delay)
                        delay *= 2
                    else:
                        raise

        # Call the retry_request function to get car details and MOT history
        return retry_request()


    def calculate_mot_score(self, mot_history):
        if mot_history == "N/A":
            return 0

        advisories_count = sum(len([comment for comment in entry['comments'] if 'ADVISORY' in comment.upper()]) for entry in mot_history)
        important_info_count = sum(len([comment for comment in entry['comments'] if 'FAIL' in comment.upper() or 'IMPORTANT' in comment.upper()]) for entry in mot_history)

        score = max(0, 100 - (advisories_count * 2 + important_info_count * 5))
        return score / 100

    def update_car_scores(self, file_path='autotrader_uk_details.xlsx'):
        df = pd.read_excel(file_path)

        options = Options()
        options.headless = False
        options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        service = Service(self.chrome_driver_path)
        driver = webdriver.Chrome(service=service, options=options)

        plates = df['Number Plate'].dropna().unique()

        results = []
        for plate in plates:
            if plate == "No number plate detected":
                mot_expiry = "N/A"
                mot_history = "N/A"
                mot_score = 0
            else:
                print(f"Fetching details for plate: {plate}")
                car_details = self.get_car_details_and_mot_history(driver, plate)
                mot_history = car_details.get('mot_history', [])
                mot_score = self.calculate_mot_score(mot_history)
                mot_expiry = car_details.get('mot_expiry', 'N/A')

            results.append((plate, mot_expiry, mot_history, mot_score))

        for plate, mot_expiry, mot_history, mot_score in results:
            df.loc[df['Number Plate'] == plate, 'MOT Expiry'] = mot_expiry
            if isinstance(mot_history, list):
                mot_history_str = "\n".join([f"Test Date: {entry['test_date']}, Mileage: {entry['mileage']}, Expiry Date: {entry['expiry_date']}, Comments: {'; '.join(entry['comments'])}" for entry in mot_history])
            else:
                mot_history_str = mot_history
            df.loc[df['Number Plate'] == plate, 'MOT History'] = mot_history_str
            df.loc[df['Number Plate'] == plate, 'MOT Score'] = mot_score

        driver.quit()

        df = self.clean_and_normalize(df)
        df.to_excel('updated_autotrader_uk_details.xlsx', index=False)

        print(f"Updated and sorted car scores saved to updated_autotrader_uk_details.xlsx")

    def clean_and_normalize(self, df):
        weights = {'Price': 0.3, 'Mileage': 0.2, 'Registration Year': 0.2, 'MOT': 0.3}

        def clean_mileage(mileage):
            try:
                return float(''.join(filter(str.isdigit, str(mileage))))
            except:
                return None

        df['Price'] = df['Price'].str.replace('£', '').str.replace(',', '').astype(float)
        df['Mileage'] = df['Mileage'].apply(clean_mileage)
        df['Registration Year'] = df['Registration Year'].str.extract(r'(\d{4})').astype(float)

        df = df.dropna(subset=['Price', 'Mileage', 'Registration Year'])

        def normalize(column):
            return (df[column] - df[column].min()) / (df[column].max() - df[column].min())

        df['Price Score'] = 1 - normalize('Price')
        df['Mileage Score'] = 1 - normalize('Mileage')
        df['Year Score'] = normalize('Registration Year')

        if 'MOT Score' not in df.columns:
            df['MOT Score'] = 0

        df['Total Score'] = (
            weights['Price'] * df['Price Score'] +
            weights['Mileage'] * df['Mileage Score'] +
            weights['Registration Year'] * df['Year Score'] +
            weights['MOT'] * df['MOT Score']
        )

        df = df.sort_values(by='Total Score', ascending=False)

        return df

    def save_links_and_update_scores(self, filename='links.txt'):
        autotrader_links = self.read_links_from_file(filename)
        self.save_links_to_spreadsheet(autotrader_links)
        print(f"Initial car details saved to autotrader_uk_details.xlsx")
        # Manual review step
        input("Press Enter after reviewing the spreadsheet to continue with updating car scores...")
        self.update_car_scores()

if __name__ == "__main__":
    scraper = AutoTraderScraper(service_account_file='service-account-file.json', chrome_driver_path=r".\chromedriver.exe")
    scraper.save_links_and_update_scores()
    scraper.update_car_scores()
